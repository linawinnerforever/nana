"""
信用卡 PDF 账单自动解析工具
适用于 Bank of America 商业信用卡 PDF 账单
部署在 Streamlit 上使用
"""

import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st


# ============================================================
# PDF 文本提取
# ============================================================

def extract_pdf_text(file_bytes):
    """多引擎容错提取 PDF 原文"""
    text = ""
    # 引擎 1: pypdf
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    except Exception:
        pass
    # 引擎 2: pdfplumber (回退)
    if not text.strip():
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
        except Exception:
            pass
    return text


# ============================================================
# PDF 数据解析
# ============================================================

# 正则: AUTO PAYMENT DEDUCTION 行 (无 MCC，参考号短)
_AUTO_PAY_RE = re.compile(
    r'(?<![/\d])(\d{2}/\d{2})\s+(?<![/\d])(\d{2}/\d{2})\s+'
    r'AUTO PAYMENT DEDUCTION\s+(\d+)\s+([\d,]+\.\d{2})'
)

# 正则: 普通交易行 PostDate TransDate Description RefNum(20+) MCC(4) Charge [Credit]
# (?<![/\d]) 防止从 "07/31/26" 中误匹配 "31/26" 为日期
# 末尾可选 Credit 金额: 若 Credit 列有数据, 金额取 Credit 的负数
_TX_RE = re.compile(
    r'(?<![/\d])(\d{2}/\d{2})\s+(?<![/\d])(\d{2}/\d{2})\s+'
    r'(.+?)\s+(\d{20,})\s+(\d{4})\s+([\d,]+\.\d{2})'
    r'(?:\s+(-?[\d,]+\.\d{2}))?'
)

# 正则: 外币换算续行  "07/03  431.28  SGD  1.279686"
_FX_RE = re.compile(
    r'(\d{2}/\d{2})\s+([\d,]+\.\d{2})\s+([A-Z]{3})\s+([\d.]+)'
)

# 正则: 乘客姓名续行  "JIA/YI"  "SUN/NIAN"
_PASSENGER_RE = re.compile(r'^[A-Z]+/[A-Z]+\s*$')

# 正则: Arrival 续行  "Arrival: 07/08/26"
_ARRIVAL_RE = re.compile(r'Arrival:\s*\d{2}/\d{2}/\d{2}')

# 正则: Cardholder Activity Summary 表格行
#   格式: 持卡人姓名 \n XXXX-XXXX-XXXX-卡号末四位 \n 数字行(5个数)
#   5个数依次为: Credit Limit, Credits, Cash, Purchases and Other Debits, Total Activity
_CARDHOLDER_SUMMARY_RE = re.compile(
    r'([A-Z][A-Z\s,.]+?)\n'
    r'XXXX-XXXX-XXXX-(\d{4})\n'
    r'([\d,]+(?:\s+[\d,]+){4})'
)

# 需要跳过的非交易行关键词
_SKIP_KEYWORDS = {
    'Total Activity', 'Account Number', 'Posting', 'Transaction',
    'Description', 'Reference Number', 'MCC', 'Charge', 'Credit',
    'Departure Date', 'Airport Code', 'Page ', 'Company Statement',
    'Finance Charge', 'Annual', 'Percentage Rate', 'PURCHASES', 'CASH',
}


def _format_continuation(line):
    """将续行文本格式化为附加描述片段，返回 None 表示跳过"""
    line = line.strip()
    if not line:
        return None

    # 跳过表头、摘要等非交易文本
    for kw in _SKIP_KEYWORDS:
        if kw in line:
            return None

    # 纯数字行 (如 booking ref "0162125281868") — 跳过
    if re.fullmatch(r'[\d\s]+', line):
        return None

    # 航空代码行 "UA S LAX" "UA W SFO" — 跳过
    if re.fullmatch(r'[A-Z]{2}\s+[A-Z]\s+[A-Z]{3}', line):
        return None

    # 外币换算行 → "(431.28 SGD)"
    m = _FX_RE.match(line)
    if m:
        return f"({m.group(2)} {m.group(3)})"

    # 乘客姓名行 → "(JIA/YI)"
    if _PASSENGER_RE.match(line):
        return f"({line})"

    # Arrival 行 → "(Arrival: 07/08/26)"
    m = _ARRIVAL_RE.search(line)
    if m:
        return f"({m.group(0)})"

    return None  # 其他续行默认跳过


def _parse_section_content(content, period, holder_name, card_last4):
    """
    逐行解析持卡人区块内容，返回 (total_activity, transactions_list)
    续行(乘客名/到达日/外币换算)会附加到前一条交易的描述中
    """
    total_activity = None
    txns = []

    lines = content.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 空行跳过
        if not line:
            i += 1
            continue

        # Total Activity 金额行
        total_m = re.match(r'Total Activity\s*(-?\$?[\d,]+\.\d{2})', line)
        if total_m:
            total_str = total_m.group(1).replace('$', '').replace(',', '')
            total_activity = float(total_str)
            i += 1
            continue

        # AUTO PAYMENT DEDUCTION 行
        m = _AUTO_PAY_RE.match(line)
        if m:
            post_date, trans_date, ref_num = m.group(1), m.group(2), m.group(3)
            amount = float(m.group(4).replace(',', ''))
            txns.append([
                period, holder_name, card_last4,
                trans_date, post_date,
                "AUTO PAYMENT DEDUCTION", ref_num, "",
                -abs(amount), "Credit / Payment"
            ])
            i += 1
            continue

        # 普通交易行
        m = _TX_RE.match(line)
        if m:
            post_date, trans_date = m.group(1), m.group(2)
            raw_desc, ref_num, mcc = m.group(3), m.group(4), m.group(5)
            charge_str = m.group(6)
            credit_str = m.group(7) if m.lastindex and m.lastindex >= 7 else None

            # Credit 列有数据 → 金额取 Credit 的负数
            if credit_str:
                amount = -abs(float(credit_str.replace(',', '')))
                tx_type = "Credit"
            else:
                amount = float(charge_str.replace(',', ''))
                tx_type = "Charge"

            desc = re.sub(r'\s+', ' ', raw_desc).strip()

            # 向后收集续行，附加到描述
            i += 1
            while i < len(lines):
                next_line = lines[i].strip()
                if not next_line:
                    i += 1
                    continue
                # 如果下一行是新的交易行 → 停止
                if _TX_RE.match(next_line) or _AUTO_PAY_RE.match(next_line):
                    break
                if next_line.startswith('Total Activity') or 'Account Number' in next_line:
                    break
                # 尝试格式化为续行片段
                frag = _format_continuation(next_line)
                if frag:
                    desc += f" {frag}"
                i += 1

            txns.append([
                period, holder_name, card_last4,
                trans_date, post_date,
                desc, ref_num, mcc,
                amount, tx_type
            ])
            continue

        i += 1

    return total_activity, txns


def parse_cardholder_summary(text, main_acct_last4):
    """
    从 PDF 的「Cardholder Activity Summary」表格提取汇总数据。
    返回 {持卡人名: Total Activity 金额} (排除公司层级)。
    表格格式:
        BHAGAT, KRUTI V
        XXXX-XXXX-XXXX-4574
        200,000 0.00 0.00 16,954.88 16,954.88   <- 5个数，最后一个为 Total Activity
    """
    summary = {}

    # 定位最后一个 Cardholder Activity Summary 表格 (实际数据表)
    idx = text.rfind('Cardholder Activity Summary')
    if idx == -1:
        return summary

    # 限定解析范围到 Transactions 之前，避免误匹配交易明细区
    tx_idx = text.find('Transactions', idx)
    segment = text[idx:tx_idx] if tx_idx > idx else text[idx:]

    for m in _CARDHOLDER_SUMMARY_RE.finditer(segment):
        holder = m.group(1).strip()
        card_last4 = m.group(2)
        numbers = [float(x.replace(',', '')) for x in m.group(3).split()]
        if len(numbers) != 5:
            continue
        total_activity = numbers[-1]
        # 排除公司层级 (卡号与主账户相同 或 名称为 CRAZY MAPLE STUDIO)
        if card_last4 != main_acct_last4 and 'CRAZY MAPLE' not in holder.upper():
            summary[holder] = total_activity

    return summary


def parse_pdf(file_bytes):
    """
    解析单个 PDF 账单，返回:
      - period: 账单周期字符串
      - summary: {持卡人名: Total Activity 金额}  (排除公司层级)
      - transactions: [[period, holder, card4, trans_date, post_date, desc, ref, mcc, amount, type], ...]
    """
    text = extract_pdf_text(file_bytes)
    if not text.strip():
        return None, {}, []

    # 1. 提取账单周期
    period_match = re.search(
        r'([A-Za-z]+\s+\d{2},\s*\d{4}\s*-\s*[A-Za-z]+\s+\d{2},\s*\d{4})',
        text
    )
    period = period_match.group(1).strip() if period_match else "Unknown Period"

    # 2. 提取主账户号末四位 (用于排除公司层级持卡人)
    acct_match = re.search(r'XXXX-XXXX-XXXX-(\d{4})', text)
    main_acct_last4 = acct_match.group(1) if acct_match else None

    # 3. 定位 Transactions 部分
    tx_start = text.find("Transactions")
    if tx_start == -1:
        return period, {}, []
    tx_text = text[tx_start:]

    # 4. 查找所有持卡人区块
    #    格式: 持卡人姓名(全大写) + \nAccount Number: XXXX-XXXX-XXXX-XXXX
    #    [A-Z\s,.] 仅匹配大写字母、空白、逗号、句点 — 不会误匹配页面页眉(含-和数字)
    section_pattern = r'([A-Z][A-Z\s,.]+?)\nAccount Number:\s*XXXX-XXXX-XXXX-(\d{4})'
    sections = list(re.finditer(section_pattern, tx_text))

    # 汇总 Dashboard 数据来源: Cardholder Activity Summary (排除公司层级)
    summary = parse_cardholder_summary(text, main_acct_last4)

    transactions = []

    for idx, section in enumerate(sections):
        holder_name = section.group(1).strip()
        card_last4 = section.group(2)

        # 获取区块内容 (从当前 header 结尾到下一个 header 开头)
        start_pos = section.end()
        end_pos = sections[idx + 1].start() if idx + 1 < len(sections) else len(tx_text)
        content = tx_text[start_pos:end_pos]

        # 逐行解析交易明细 (汇总不再从这里取)
        _, txns = _parse_section_content(
            content, period, holder_name, card_last4
        )

        transactions.extend(txns)

    # 兜底: 若 Cardholder Activity Summary 解析失败, 退回从交易明细区 Total Activity 提取
    if not summary:
        for idx, section in enumerate(sections):
            holder_name = section.group(1).strip()
            card_last4 = section.group(2)
            start_pos = section.end()
            end_pos = sections[idx + 1].start() if idx + 1 < len(sections) else len(tx_text)
            content = tx_text[start_pos:end_pos]
            total_m = re.search(r'Total Activity\s*(-?\$?[\d,]+\.\d{2})', content)
            if total_m:
                total_str = total_m.group(1).replace('$', '').replace(',', '')
                total_amount = float(total_str)
                if card_last4 != main_acct_last4 and 'CRAZY MAPLE' not in holder_name:
                    summary[holder_name] = total_amount

    return period, summary, transactions


# ============================================================
# Excel 报表生成
# ============================================================

def generate_excel(summary_by_period, transactions):
    """生成与参考格式一致的 Excel 汇总表 (BytesIO)"""
    wb = openpyxl.Workbook()

    # --- 样式定义 ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    stat_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    total_border = Border(
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="double", color="000000"),
        left=thin_side, right=thin_side
    )

    # ============================================================
    # Sheet 1: 交易明细
    # ============================================================
    ws_det = wb.create_sheet(title="交易明细")
    ws_det.views.sheetView[0].showGridLines = True

    # 前两行留空 (I1/I2 存放校验公式)
    for r in [1, 2]:
        for c in range(1, 11):
            cell = ws_det.cell(row=r, column=c, value="")
            cell.fill = stat_fill
            cell.border = thin_border

    last_row = len(transactions) + 3

    # I1: 公司层级 (CRAZY MAPLE STUDIO) 交易合计
    c = ws_det.cell(
        row=1, column=9,
        value=f'=SUMIF(B4:B{last_row}, "CRAZY MAPLE STUDIO*", I4:I{last_row})'
    )
    c.font = bold_font
    c.fill = stat_fill
    c.number_format = "$#,##0.00"
    c.alignment = Alignment(horizontal="right", vertical="center")

    # I2: 非公司持卡人 Charge 合计
    c = ws_det.cell(
        row=2, column=9,
        value=f'=SUMIFS(I4:I{last_row}, B4:B{last_row}, "<>CRAZY MAPLE STUDIO*", J4:J{last_row}, "Charge")'
    )
    c.font = bold_font
    c.fill = stat_fill
    c.number_format = "$#,##0.00"
    c.alignment = Alignment(horizontal="right", vertical="center")

    # 第 3 行: 表头
    det_headers = [
        "账单周期", "持卡人 / 账户", "卡号末四位",
        "交易日 (Trans Date)", "记账日 (Post Date)",
        "交易描述 (Description)", "参考号 (Reference No)",
        "MCC", "金额 (Amount)", "交易类型 (Type)"
    ]
    for i, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=3, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 数据行 (第 4 行起)
    for r_idx, row_data in enumerate(transactions, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_det.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c_idx in [1, 3, 4, 5, 8, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [2, 6, 7]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 9:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # 自动筛选 + 冻结前 3 行
    ws_det.auto_filter.ref = f"A3:J{last_row}"
    ws_det.freeze_panes = "A4"

    # 自适应列宽
    for col_idx in range(1, 11):
        letter = get_column_letter(col_idx)
        max_len = len(det_headers[col_idx - 1])
        for r in range(4, last_row + 1):
            v = ws_det.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        if col_idx == 9:
            ws_det.column_dimensions[letter].width = max(max_len + 5, 18)
        elif col_idx == 6:
            ws_det.column_dimensions[letter].width = min(max_len + 4, 55)
        else:
            ws_det.column_dimensions[letter].width = max(max_len + 4, 14)

    # ============================================================
    # Sheet 2: 汇总 Dashboard
    # ============================================================
    ws_sum = wb.active
    ws_sum.title = "汇总 Dashboard"
    ws_sum.views.sheetView[0].showGridLines = True

    periods = list(summary_by_period.keys())
    sum_headers = ["持卡人"] + periods + ["合计 (Total)"]

    for i, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    all_holders = sorted({h for p in summary_by_period.values() for h in p.keys()})

    for r_idx, holder in enumerate(all_holders, start=2):
        c = ws_sum.cell(row=r_idx, column=1, value=holder)
        c.font = data_font
        c.border = thin_border

        row_total = 0.0
        for p_idx, p_name in enumerate(periods, start=2):
            val = summary_by_period[p_name].get(holder, 0.0)
            row_total += val
            c = ws_sum.cell(row=r_idx, column=p_idx, value=val)
            c.font = data_font
            c.number_format = "$#,##0.00"
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thin_border

        c = ws_sum.cell(row=r_idx, column=len(periods) + 2, value=row_total)
        c.font = data_font
        c.number_format = "$#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = thin_border

    # 合计行
    tot_row = len(all_holders) + 2
    lbl = ws_sum.cell(row=tot_row, column=1, value="合计 (Total)")
    lbl.font = bold_font
    lbl.border = total_border

    grand_total = 0.0
    for p_idx, p_name in enumerate(periods, start=2):
        col_total = sum(summary_by_period[p_name].values())
        grand_total += col_total
        c = ws_sum.cell(row=tot_row, column=p_idx, value=col_total)
        c.font = bold_font
        c.number_format = "$#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = total_border

    total_letter = get_column_letter(len(periods) + 2)
    c = ws_sum.cell(row=tot_row, column=len(periods) + 2, value=grand_total)
    c.font = bold_font
    c.number_format = "$#,##0.00"
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = total_border

    # Check 校验行: 交易明细中非公司 Charge 合计 - 汇总表总计 (应为 0)
    check_row = tot_row + 2
    ws_sum.cell(row=check_row, column=1, value="check").font = bold_font
    c = ws_sum.cell(row=check_row, column=len(periods) + 2)
    c.value = f'=交易明细!I2 - {total_letter}{tot_row}'
    c.font = bold_font
    c.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    c.alignment = Alignment(horizontal="right", vertical="center")

    # 列宽
    ws_sum.column_dimensions['A'].width = 25
    for col_i in range(2, len(periods) + 3):
        ws_sum.column_dimensions[get_column_letter(col_i)].width = 30

    # 输出
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="信用卡账单自动解析", page_icon="📄")
st.title("📄 信用卡 PDF 账单自动解析工具")
st.write("上传 Bank of America 信用卡 PDF 账单（支持多文件），自动生成汇总表。")

uploaded_files = st.file_uploader("选择 PDF 文件", type=["pdf"], accept_multiple_files=True)

if uploaded_files:
    if st.button("🚀 开始提取并生成 Excel", type="primary"):
        with st.spinner("正在解析 PDF 并生成报表..."):
            try:
                summary_by_period = {}
                all_transactions = []

                for f in uploaded_files:
                    period, summary, txns = parse_pdf(f.getvalue())
                    if period is None:
                        continue
                    if period not in summary_by_period:
                        summary_by_period[period] = {}
                    summary_by_period[period].update(summary)
                    all_transactions.extend(txns)

                if not all_transactions:
                    st.warning("未能从 PDF 中提取任何交易明细，请检查文件格式是否为 BofA 信用卡账单。")
                else:
                    excel_bytes = generate_excel(summary_by_period, all_transactions)
                    st.success(
                        f"成功解析 {len(uploaded_files)} 个 PDF 账单！"
                        f"已提取 {len(all_transactions)} 条交易明细。"
                    )
                    st.download_button(
                        label="📥 下载 Excel 汇总表格",
                        data=excel_bytes,
                        file_name="信用卡账单汇总表.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"解析过程出错: {e}")
