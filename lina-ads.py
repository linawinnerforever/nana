import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import re

st.set_page_config(page_title="投放费用数据智能汇总工具", layout="wide")

st.title("📊 投放费用自动入账工具")
st.markdown("特性：上传业务底表，自动输出汇总表&入账凭证")

# 提供双文件上传器
col_u1, col_u2 = st.columns(2)
with col_u1:
    uploaded_files = st.file_uploader("1. 上传业务计提表 (可多选 Excel)", type=["xlsx", "xls"], accept_multiple_files=True)
with col_u2:
    mp_file = st.file_uploader("2. 上传最新的投放费用 MP 主数据映射表 (单选 Excel)", type=["xlsx", "xls"])

PRODUCT_MAPPING = {
    'chapters': {'product': 'CHAPTERS', 'entity': 'CM'},
    'kiss': {'product': 'KISS', 'entity': 'MH'},
    'maxdrama': {'product': 'MaxDrama', 'entity': 'NL'}, 
    'merge': {'product': 'Merge', 'entity': 'CM'},
    'reelshort': {'product': 'Reelshort', 'entity': 'NL'},
    'rsnovel': {'product': 'RS N', 'entity': 'NL'}
}

REQUIRED_SHEETS = {
    'Advertising-Chapters': {'key': 'chapters', 'product': 'CHAPTERS', 'entity': 'CM'},
    'Advertising-Kiss': {'key': 'kiss', 'product': 'KISS', 'entity': 'MH'},
    'Advertising-MaxDrama': {'key': 'maxdrama', 'product': 'MaxDrama', 'entity': 'NL'},
    'Advertising-Merge': {'key': 'merge', 'product': 'Merge', 'entity': 'CM'},
    'Advertising-Reelshort': {'key': 'reelshort', 'product': 'Reelshort', 'entity': 'NL'},
    'Advertising-RS N': {'key': 'rsnovel', 'product': 'RS N', 'entity': 'NL'}
}

# 严格锁定 71 列金蝶凭证表头矩阵映射关系 (物理列 1-71)
STRICT_71_VOUCHER_HEADERS = [
    ("FBillHead(GL_VOUCHER)", "*单据头(序号)"), ("FAccountBookID", "*(单据头)账簿#编码"), ("FAccountBookID#Name", "(单据头)账簿#名称"),
    ("FDate", "*(单据头)日期"), ("FBUSDATE", "(单据头)业务日期"), ("FYEAR", "(单据头)会计年度"), ("FPERIOD", "(单据头)期间"),
    ("FVOUCHERGROUPID", "*(单据头)凭证字#编码"), ("FVOUCHERGROUPID#Name", "(单据头)凭证字#名称"), ("FVOUCHERGROUPNO", "*(单据头)凭证号"),
    ("FATTACHMENTS", "(单据头)附件数"), ("FISADJUSTVOUCHER", "(单据头)是否调整期凭证"), ("FACCBOOKORGID", "*(单据头)核算组织#编码"),
    ("FACCBOOKORGID#Name", "(单据头)核算组织#名称"), ("FSourceBillKey", "(单据头)业务类型#编码"), ("FSourceBillKey#Name", "(单据头)业务类型#名称"),
    ("FIMPORTVERSION", "(单据头)引入版本号"), ("*Split*1", "间隔列"), ("FEntity", "*分录(序号)"), ("FEXPLANATION", "(分录)摘要"),
    ("FACCOUNTID", "*(分录)科目编码#编码"), ("FACCOUNTID#Name", "(分录)科目编码#名称"), ("FDetailID#FF100003", "(分录)北京公司项目名#编码"),
    ("FDetailID#FF100003#Name", "(分录)北京公司项目名#名称(Null)"), ("FDetailID#FF100002", "(分录)项目段#编码"),
    ("FDetailID#FF100002#Name", "(分录)项目段#名称(Null)"), ("FDetailID#FFLEX16", "(分录)其他往来单位#编码"),
    ("FDetailID#FFLEX16#Name", "(分录)其他往来单位#名称(Null)"), ("FDetailID#FFLEX15", "(分录)银行账号#编码"),
    ("FDetailID#FFLEX15#Name", "(分录)银行账号#名称(Null)"), ("FDetailID#FFLEX14", "(分录)银行#编码"),
    ("FDetailID#FFLEX14#Name", "(分录)银行#名称(Null)"), ("FDetailID#FFLEX13", "(分录)客户分组#编码"),
    ("FDetailID#FFLEX13#Name", "(分录)客户分组#名称(Null)"), ("FDetailID#FFLEX12", "(分录)物料分组#编码"),
    ("FDetailID#FFLEX12#Name", "(分录)物料分组#名称(Null)"), ("FDetailID#FFLEX11", "(分录)组织机构#编码"),
    ("FDetailID#FFLEX11#Name", "(分录)组织机构#名称(Null)"), ("FDetailID#FFlex10", "(分录)资产类别#编码"),
    ("FDetailID#FFlex10#Name", "(分录)资产类别#名称(Null)"), ("FDetailID#FFLEX9", "(分录)费用项目#编码"),
    ("FDetailID#FFLEX9#Name", "(分录)费用项目#名称(Null)"), ("FDetailID#FFlex8", "(分录)物料#编码"),
    ("FDetailID#FFlex8#Name", "(分录)物料#名称(Null)"), ("FDetailID#FFlex7", "(分录)员工#编码"),
    ("FDetailID#FFlex7#Name", "(分录)员工#名称(Null)"), ("FDetailID#FFlex6", "(分录)客户#编码"),
    ("FDetailID#FFlex6#Name", "(分录)客户#名称(Null)"), ("FDetailID#FFlex5", "(分录)部门#编码"),
    ("FDetailID#FFlex5#Name", "(分录)部门#名称(Null)"), ("FDetailID#FFlex4", "(分录)供应商#编码"),
    ("FDetailID#FFlex4#Name", "(分录)供应商#名称(Null)"), ("FDetailID#FF100004", "(分录)NL剧集#编码"),
    ("FDetailID#FF100004#Name", "(分录)NL剧集#名称(Null)"), ("FDetailID#FF100005", "(分录)海南剧集#编码"),
    ("FDetailID#FF100005#Name", "(分录)海南剧集#名称(Null)"), ("FCURRENCYID", "*(分录)币别#编码"),
    ("FCURRENCYID#Name", "(分录)币别#名称"), ("FEXCHANGERATETYPE", "*(分录)汇率类型#编码"), ("FEXCHANGERATETYPE#Name", "(分录)汇率类型#名称"),
    ("FEXCHANGERATE", "(分录)汇率"), ("FUnitId", "(分录)单位#编码"), ("FUnitId#Name", "(分录)单位#名称"),
    ("FPrice", "(分录)单价"), ("FQty", "(分录)数量"), ("FAMOUNTFOR", "(分录)原币金额"), ("FDEBIT", "(分录)借方金额"),
    ("FCREDIT", "(分录)贷方金额"), ("FSettleTypeID", "(分录)结算方式#编码"), ("FSettleTypeID#Name", "(分录)结算方式#名称"),
    ("FSETTLENO", "(分录)结算号"), ("FBUSNO", "(分录)业务编号"), ("FEXPORTENTRYID", "(分录)现金流量#分录ID")
]

def clean_amount(val):
    if pd.isna(val):
        return 0.0
    val_str = str(val).strip().replace('$', '').replace(',', '')
    val_str = val_str.replace('−', '-').replace('—', '-')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

@st.cache_data
def load_mp_matrix(file):
    try:
        df_mp = pd.read_excel(file, skiprows=1, dtype=str)
        df_mp.columns = [str(c).strip() for c in df_mp.columns]
        return df_mp
    except Exception as e:
        st.error(f"解析 MP 映射表失败: {str(e)}")
        return None

def build_openpyxl_voucher_strict_71(df_source, entity_name, month_str):
    df_filter = df_source[(df_source['核算主体'] == entity_name) & (df_source['供应商-金蝶'] != "")].copy()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "凭证#单据头(FBillHead)"
    ws.views.sheetView[0].showGridLines = True
    
    for col_idx, (en_h, cn_h) in enumerate(STRICT_71_VOUCHER_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=en_h)
        ws.cell(row=2, column=col_idx, value=cn_h)
        
    en_months = {"1": "Jan", "2": "Feb", "3": "Mar", "4": "Apr", "5": "May", "6": "Jun", "7": "Jul", "8": "Aug", "9": "Sep", "10": "Oct", "11": "Nov", "12": "Dec"}
    en_month_label = en_months.get(month_str, "Month")
    
    book_id = "002" if entity_name == "CM" else "005"
    org_id = "100" if entity_name == "CM" else "103"
    
    df_grouped = df_filter.groupby(['项目', '投放渠道', '开户方', '供应商-金蝶', '供应商编码'], dropna=False)
    
    current_row = 3
    
    for idx, (group_keys, group_data) in enumerate(df_grouped, start=1):
        p_project, p_channel, p_kf, p_jindie, p_code = group_keys
        p_spent = group_data['spent'].sum()
        if p_spent <= 0:
            continue
            
        if str(p_kf).strip() != "":
            explanation = f"计提2026年{month_str}月推广费用- {en_month_label}.2026 advertising and marketing cost accrual-{p_channel}-{p_kf}"
        else:
            explanation = f"计提2026年{month_str}月推广费用- {en_month_label}.2026 advertising and marketing cost accrual-{p_channel}"
            
        # ----------------------------------------------------
        # 一、借方分录行填充 (6601.03.01)
        # ----------------------------------------------------
        if idx == 1:
            ws.cell(row=current_row, column=1, value=1)  # FBillHead(GL_VOUCHER)
            ws.cell(row=current_row, column=2, value=book_id)  # FAccountBookID
            ws.cell(row=current_row, column=4, value=f"2026-{month_str.zfill(2)}-30")  # FDate
            ws.cell(row=current_row, column=5, value=f"2026-{month_str.zfill(2)}-30")  # FBUSDATE
            ws.cell(row=current_row, column=6, value=2026)  # FYEAR
            ws.cell(row=current_row, column=7, value=int(month_str))  # FPERIOD
            ws.cell(row=current_row, column=8, value="PRE001")  # FVOUCHERGROUPID
            ws.cell(row=current_row, column=10, value=1)  # FVOUCHERGROUPNO
            ws.cell(row=current_row, column=13, value=org_id)  # FACCBOOKORGID
            
        ws.cell(row=current_row, column=19, value=(idx - 1) * 2 + 1)  # FEntity
        ws.cell(row=current_row, column=20, value=explanation)  # FEXPLANATION
        ws.cell(row=current_row, column=21, value="6601.03.01")  # FACCOUNTID
        
        # 【精准修正点】：项目段编码依照财务最新指示移回 FDetailID#FF100002（第 25 列）
        ws.cell(row=current_row, column=25, value=p_project)  
        
        ws.cell(row=current_row, column=49, value="70000")  # FDetailID#FFlex5 (第49列：项目#编码 7000)
        ws.cell(row=current_row, column=57, value="PRE007")  # FCURRENCYID
        ws.cell(row=current_row, column=58, value="美元")  # FCURRENCYID#Name
        ws.cell(row=current_row, column=59, value="HLTX01_SYS")  # FEXCHANGERATETYPE
        ws.cell(row=current_row, column=60, value="固定汇率")  # FEXCHANGERATETYPE#Name
        ws.cell(row=current_row, column=61, value=1)  # FEXCHANGERATE
        
        # 借方纯数字千分位
        cell_dr = ws.cell(row=current_row, column=67, value=round(p_spent, 2)) # 第67列 FDEBIT
        cell_dr.number_format = '#,##0.00'
        
        current_row += 1
        
        # ----------------------------------------------------
        # 二、贷方分录行填充 (2202.02)
        # ----------------------------------------------------
        ws.cell(row=current_row, column=19, value=(idx - 1) * 2 + 2)  # FEntity
        ws.cell(row=current_row, column=20, value=explanation)  # FEXPLANATION
        ws.cell(row=current_row, column=21, value="2202.02")  # FACCOUNTID
        ws.cell(row=current_row, column=51, value=p_code)  # FDetailID#FFlex4 (第51列：供应商#编码)
        ws.cell(row=current_row, column=57, value="PRE007")  # FCURRENCYID
        ws.cell(row=current_row, column=58, value="美元")  # FCURRENCYID#Name
        ws.cell(row=current_row, column=59, value="HLTX01_SYS")  # FEXCHANGERATETYPE
        ws.cell(row=current_row, column=60, value="固定汇率")  # FEXCHANGERATETYPE#Name
        ws.cell(row=current_row, column=61, value=1)  # FEXCHANGERATE
        
        # 贷方纯数字千分位
        cell_cr = ws.cell(row=current_row, column=68, value=round(p_spent, 2)) # 第68列 FCREDIT
        cell_cr.number_format = '#,##0.00'
        
        current_row += 1
        
    # 财务前置文本格式保护设定，锁死文本格式防篡改
    for r in range(3, current_row):
        ws.cell(row=r, column=2).number_format = '@'
        ws.cell(row=r, column=13).number_format = '@'
        ws.cell(row=r, column=25).number_format = '@' # 第25列项目文本
        ws.cell(row=r, column=49).number_format = '@'
        ws.cell(row=r, column=51).number_format = '@'
        
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

def process_data_flow(files):
    all_data = []
    detected_period = "未知期间"
    file_month_label = "X月"
    
    for f in files:
        match = re.search(r'(\d+年\d+月|\d+月)', f.name)
        if match:
            raw_month = match.group(1)
            if "年" in raw_month:
                year_part = raw_month.split("年")[0][-2:]
                month_part = raw_month.split("年")[1]
                detected_period = f"{month_part}-{year_part}" 
                file_month_label = month_part
            else:
                detected_period = f"{raw_month}-26"
                file_month_label = raw_month
            break

    for f in files:
        fname = f.name.lower()
        prod_info = None
        target_sheet_name = None
        
        for s_name, info in REQUIRED_SHEETS.items():
            if info['key'] in fname:
                prod_info = info
                target_sheet_name = s_name
                break
        if not prod_info:
            continue
            
        try:
            df_check = pd.read_excel(f, nrows=10)
            header_row = 0
            for i, row in df_check.iterrows():
                if '投放渠道' in row.values or 'spent' in row.values or '消耗' in row.values:
                    header_row = i + 1
                    break
            
            df = pd.read_excel(f, skiprows=header_row)
            df.columns = [str(c).strip() for c in df.columns]
            
            amt_col = 'spent' if 'spent' in df.columns else ('消耗' if '消耗' in df.columns else None)
            if not amt_col:
                continue
                
            df = df[df[amt_col].notna()]
            df['spent_num'] = df[amt_col].apply(clean_amount)
            
            channel_col = '投放渠道' if '投放渠道' in df.columns else df.columns[0]
            df = df[~df[channel_col].astype(str).str.contains('合计|Total', na=False)]
            
            processed_df = pd.DataFrame()
            processed_df['期间'] = detected_period
            processed_df['投放渠道'] = df['投放渠道'].fillna('').astype(str).str.strip()
            
            if '开户方' in df.columns:
                processed_df['开户服务商'] = df['开户方'].fillna('').astype(str).str.strip()
            elif '开户服务商' in df.columns:
                processed_df['开户服务商'] = df['开户服务商'].fillna('').astype(str).str.strip()
            else:
                processed_df['开户服务商'] = ''
                
            processed_df['广告户名'] = df['广告户名'].fillna('').astype(str).str.strip() if '广告户名' in df.columns else ''
            processed_df['类别'] = df['类别'].fillna('自投').astype(str).str.strip() if '类别' in df.columns else '自投'
            processed_df['代投服务商'] = df['代投服务商'].fillna('').astype(str).str.strip() if '代投服务商' in df.columns else ''
            
            processed_df['消耗'] = df['spent_num']
            processed_df['代投费'] = df['代投费'].apply(clean_amount) if '代投费' in df.columns else 0.0
            processed_df['投放待结算'] = processed_df['消耗'] + processed_df['代投费']
            
            processed_df['买量产品'] = prod_info['product']
            processed_df['核算主体'] = prod_info['entity']
            processed_df['Target_Sheet'] = target_sheet_name
            
            all_data.append(processed_df)
        except Exception as e:
            st.error(f"处理文件 {f.name} 时出错: {str(e)}")
            
    if all_data:
        full_df = pd.concat(all_data, ignore_index=True)
        return full_df, detected_period, file_month_label
    return None, detected_period, file_month_label

if uploaded_files:
    with st.spinner("正在智能清洗流水明细数据..."):
        df_detail, current_month, month_label = process_data_flow(uploaded_files)
        
    if df_detail is not None:
        df_mp_matrix = None
        if mp_file:
            df_mp_matrix = load_mp_matrix(mp_file)
            
        df_pivot = df_detail.groupby(['买量产品', '核算主体', '投放渠道', '开户服务商'], as_index=False, dropna=False)['消耗'].sum()
        df_pivot.rename(columns={'开户服务商': '开户方', '消耗': 'spent'}, inplace=True)
        
        df_pivot['供应商-渠道'] = df_pivot.apply(
            lambda r: str(r['开户方']).strip() if str(r['开户方']).strip() != "" else str(r['投放渠道']).strip(), axis=1
        )
        
        df_pivot['项目'] = ""
        df_pivot['供应商-金蝶'] = ""
        df_pivot['供应商编码'] = ""
        df_pivot = df_pivot.astype({'项目': 'object', '供应商-金蝶': 'object', '供应商编码': 'object'})
        
        if df_mp_matrix is not None:
            dict_project = {}
            for _, r in df_mp_matrix.iterrows():
                k_prod = str(r.iloc[13]).strip().lower() 
                v_code = str(r.iloc[14]).strip()        
                if k_prod and k_prod != 'nan':
                    dict_project[k_prod] = v_code.split('.')[0].zfill(3)
            
            dict_supplier_jindie = {}
            for _, r in df_mp_matrix.iterrows():
                k_chan = str(r.iloc[10]).strip().lower() 
                v_jindie = str(r.iloc[11]).strip()       
                if k_chan and k_chan != 'nan':
                    dict_supplier_jindie[k_chan] = v_jindie
            
            dict_cm_code_by_name = {}  
            dict_mh_code_by_name = {}  
            
            for _, r in df_mp_matrix.iterrows():
                cm_name = str(r.iloc[2]).strip()   
                cm_code = str(r.iloc[0]).strip()   
                mh_name = str(r.iloc[7]).strip()   
                mh_code = str(r.iloc[5]).strip()   
                
                if cm_name and cm_name != 'nan':
                    dict_cm_code_by_name[cm_name.lower()] = cm_code
                if mh_name and mh_name != 'nan':
                    dict_mh_code_by_name[mh_name.lower()] = mh_code

            p_chan_lower = df_pivot['供应商-渠道'].astype(str).str.strip().str.lower()
            p_prod_lower = df_pivot['买量产品'].astype(str).str.strip().str.lower()
            
            mapped_project = p_prod_lower.map(dict_project).fillna("")
            mapped_jindie = p_chan_lower.map(dict_supplier_jindie).fillna("")
            
            for idx, row in df_pivot.iterrows():
                entity = str(row['核算主体']).upper().strip()
                if entity == 'NL':
                    continue
                
                df_pivot.at[idx, '项目'] = mapped_project.iloc[idx]
                current_jindie_name = mapped_jindie.iloc[idx]
                df_pivot.at[idx, '供应商-金蝶'] = current_jindie_name
                
                if current_jindie_name != "":
                    j_key = current_jindie_name.lower().strip()
                    if entity == 'CM':
                        df_pivot.at[idx, '供应商编码'] = dict_cm_code_by_name.get(j_key, "")
                    elif entity == 'MH':
                        df_pivot.at[idx, '供应商编码'] = dict_mh_code_by_name.get(j_key, "")
                        
        pivot_cols = ['买量产品', '核算主体', 'spent', '投放渠道', '开户方', '项目', '供应商-渠道', '供应商-金蝶', '供应商编码']
        df_pivot = df_pivot[pivot_cols]
        
        # 智能漏单提示
        df_failed_check = df_pivot[(df_pivot['核算主体'] != 'NL') & ((df_pivot['供应商-金蝶'] == "") | (df_pivot['供应商编码'] == ""))]
        if not df_failed_check.empty:
            failed_channels = df_failed_check['供应商-渠道'].unique()
            st.warning(f"⚠️ **主数据匹配警报**：发现有 **{len(failed_channels)}** 个【供应商-渠道】未识别到金蝶信息，请维护 MP 映射表。")
            st.code("\n".join([f"• {chan}" for chan in failed_channels]), language="text")
        else:
            st.success("✅ **主数据核对通过**：CM/MH 主体的所有供应商名称与编码已 100% 精准匹配成功！")

        # ==========================================
        # 1. 常规业务分析总表
        # ==========================================
        wb_orig = openpyxl.Workbook(); ws_orig = wb_orig.active; ws_orig.title = "费用汇总及透视表"; ws_orig.views.sheetView[0].showGridLines = True
        font_title = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        font_header = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        font_total = Font(name="微软雅黑", size=11, bold=True, color="D32F2F"); font_body = Font(name="微软雅黑", size=10)
        fill_detail_title = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid"); fill_pivot_title = PatternFill(start_color="1E6B52", end_color="1E6B52", fill_type="solid")
        fill_detail_hdr = PatternFill(start_color="4A7BB0", end_color="4A7BB0", fill_type="solid"); fill_pivot_hdr = PatternFill(start_color="339977", end_color="339977", fill_type="solid")
        fill_zebra = PatternFill(start_color="F9FBFC", end_color="F9FBFC", fill_type="solid")
        thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
        total_border = Border(top=Side(style='medium', color='D32F2F'), bottom=Side(style='medium', color='D32F2F'), left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'))
        
        detail_cols = ['投放渠道', '开户方', '广告户名', 'spent', '买量产品', '核算主体']
        detail_end = len(df_detail) + 3; pivot_end = len(df_pivot) + 3
        
        ws_orig.cell(row=1, column=1, value="总计 (SUBTOTAL)").font = font_total
        for c in range(1, 7):
            cell = ws_orig.cell(row=1, column=c); cell.border = total_border
            if detail_cols[c-1] == 'spent':
                cell.value = f"=SUBTOTAL(9, D4:D{detail_end})"; cell.font = font_total; cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c == 1: cell.alignment = Alignment(horizontal="left", vertical="center")
            else: cell.alignment = Alignment(horizontal="center", vertical="center")
                
        ws_orig.cell(row=1, column=8, value="总计 (SUBTOTAL)").font = font_total; ws_orig.cell(row=1, column=8).alignment = Alignment(horizontal="left", vertical="center")
        cell_orig_sub = ws_orig.cell(row=1, column=10); cell_orig_sub.value = f"=SUBTOTAL(9, J4:J{pivot_end})"; cell_orig_sub.font = font_total; cell_orig_sub.number_format = '#,##0.00'; cell_orig_sub.alignment = Alignment(horizontal="right", vertical="center")
        
        ws_orig.cell(row=1, column=11, value="CM:").font = font_total; ws_orig.cell(row=1, column=11).alignment = Alignment(horizontal="right", vertical="center")
        cell_cm_sumif = ws_orig.cell(row=1, column=12, value=f'=SUMIF(I4:I{pivot_end}, \"CM\", J4:J{pivot_end})'); cell_cm_sumif.font = font_total; cell_cm_sumif.number_format = '#,##0.00'; cell_cm_sumif.alignment = Alignment(horizontal="right")
        ws_orig.cell(row=1, column=13, value="MH:").font = font_total; ws_orig.cell(row=1, column=13).alignment = Alignment(horizontal="right", vertical="center")
        cell_mh_sumif = ws_orig.cell(row=1, column=14, value=f'=SUMIF(I4:I{pivot_end}, \"MH\", J4:J{pivot_end})'); cell_mh_sumif.font = font_total; cell_mh_sumif.number_format = '#,##0.00'; cell_mh_sumif.alignment = Alignment(horizontal="right")
        ws_orig.cell(row=1, column=15, value="CM+MH:").font = font_total; ws_orig.cell(row=1, column=15).alignment = Alignment(horizontal="right", vertical="center")
        cell_cm_mh_ttl = ws_orig.cell(row=1, column=16, value=f'=L1+N1'); cell_cm_mh_ttl.font = font_total; cell_cm_mh_ttl.number_format = '#,##0.00'; cell_cm_mh_ttl.alignment = Alignment(horizontal="right")
        
        for c in range(8, 8 + len(pivot_cols)): ws_orig.cell(row=1, column=c).border = total_border
        
        ws_orig.merge_cells("A2:F2"); ws_orig["A2"] = "投放费用明细表"; ws_orig["A2"].font = font_title; ws_orig["A2"].fill = fill_detail_title; ws_orig["A2"].alignment = Alignment(horizontal="center")
        end_letter = openpyxl.utils.get_column_letter(7 + len(pivot_cols))
        ws_orig.merge_cells(f"H2:{end_letter}2"); ws_orig["H2"] = "投放费用透视表"; ws_orig["H2"].font = font_title; ws_orig["H2"].fill = fill_pivot_title; ws_orig["H2"].alignment = Alignment(horizontal="center")
        
        for idx, col in enumerate(detail_cols, 1):
            cell = ws_orig.cell(row=3, column=idx, value=col); cell.font = font_header; cell.fill = fill_detail_hdr; cell.alignment = Alignment(horizontal="center")
        for idx, col in enumerate(pivot_cols, 8):
            cell = ws_orig.cell(row=3, column=idx, value=col); cell.font = font_header; cell.fill = fill_pivot_hdr; cell.alignment = Alignment(horizontal="center")
            
        for r_idx, row in enumerate(dataframe_to_rows(df_detail.rename(columns={'开户服务商': '开户方', '消耗': 'spent'})[detail_cols], index=False, header=False), start=4):
            for c_idx, val in enumerate(row, start=1):
                cell = ws_orig.cell(row=r_idx, column=c_idx, value=val); cell.font = font_body; cell.border = thin_border
                if detail_cols[c_idx-1] == 'spent': cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
                else: cell.alignment = Alignment(horizontal="center")
                if r_idx % 2 == 0: cell.fill = fill_zebra
                    
        for r_idx, row in enumerate(dataframe_to_rows(df_pivot[pivot_cols], index=False, header=False), start=4):
            for c_idx, val in enumerate(row, start=8):
                cell = ws_orig.cell(row=r_idx, column=c_idx, value=val); cell.font = font_body; cell.border = thin_border
                if pivot_cols[c_idx-8] == 'spent': cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center")
                    if pivot_cols[c_idx-8] in ['项目', '供应商编码']: cell.number_format = '@'

        ws_orig.row_dimensions[1].height = 26; ws_orig.row_dimensions[2].height = 24; ws_orig.row_dimensions[3].height = 22
        for col in ws_orig.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_orig.column_dimensions[col_letter].width = max(max_len + 3, 12)
        ws_orig.column_dimensions['G'].width = 4
        
        excel_data_orig = io.BytesIO(); wb_orig.save(excel_data_orig); excel_data_orig.seek(0)

        # ==========================================
        # 2. 导出：【给领导的汇总表】
        # ==========================================
        wb_leader = openpyxl.Workbook(); wb_leader.remove(wb_leader.active)
        font_l_hdr = Font(name="微软雅黑", size=10, bold=True); font_l_body = Font(name="微软雅黑", size=10)
        font_l_top_sub = Font(name="Arial", size=11, bold=True); font_l_grand_total = Font(name="Arial", size=11, bold=True, color="FF0000") 
        align_center = Alignment(horizontal="center", vertical="center"); align_right = Alignment(horizontal="right", vertical="center")
        leader_thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
        leader_headers = ["期间", "投放渠道", "开户服务商", "广告户名", "类别", "代投服务商", "消耗", "代投费", "投放待结算", "买量产品", "核算主体"]
        
        for sheet_name in REQUIRED_SHEETS.keys():
            ws_l = wb_leader.create_sheet(title=sheet_name); ws_l.views.sheetView[0].showGridLines = True
            df_sub = df_detail[df_detail['Target_Sheet'] == sheet_name] if 'Target_Sheet' in df_detail.columns else pd.DataFrame()
            if df_sub.empty:
                df_l_final = pd.DataFrame(columns=leader_headers)
                empty_row = {h: "" for h in leader_headers}; empty_row['期间'] = current_month; empty_row['买量产品'] = REQUIRED_SHEETS[sheet_name]['product']; empty_row['核算主体'] = REQUIRED_SHEETS[sheet_name]['entity']
                df_l_final = pd.DataFrame([empty_row])
            else:
                df_l_final = df_sub[leader_headers].copy(); df_l_final['期间'] = current_month
                
            data_end_row = max(4, len(df_l_final) + 3)
            if sheet_name == 'Advertising-Chapters':
                ws_l.cell(row=1, column=5, value="TTL:").font = font_l_hdr; ws_l.cell(row=1, column=5).alignment = align_right
                ws_l.cell(row=1, column=6, value=f"=SUM('Advertising-Chapters:Advertising-RS N'!G4:G5000)").font = font_l_grand_total; ws_l.cell(row=1, column=6).number_format = '#,##0.00'; ws_l.cell(row=1, column=6).alignment = align_right
            
            ws_l.cell(row=1, column=7, value=f"=SUM(G4:G{data_end_row})").font = font_l_top_sub; ws_l.cell(row=1, column=7).number_format = '#,##0.00'; ws_l.cell(row=1, column=7).alignment = align_right
            ws_l.cell(row=1, column=9, value=f"=SUM(I4:I{data_end_row})").font = font_l_top_sub; ws_l.cell(row=1, column=9).number_format = '#,##0.00'; ws_l.cell(row=1, column=9).alignment = align_right
            
            for idx, h_name in enumerate(leader_headers, 1):
                cell = ws_l.cell(row=3, column=idx, value=h_name); cell.font = font_l_hdr; cell.alignment = align_center; cell.border = leader_thin_border
            ws_l.row_dimensions[3].height = 24
            for r_idx, row in enumerate(dataframe_to_rows(df_l_final, index=False, header=False), start=4):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws_l.cell(row=r_idx, column=c_idx, value=val); cell.font = font_l_body; cell.border = leader_thin_border
                    if leader_headers[c_idx-1] in ["消耗", "代投费", "投放待结算"]:
                        if val != "": cell.number_format = '#,##0.00'
                        cell.alignment = align_right
                    else: cell.alignment = align_center
                ws_l.row_dimensions[r_idx].height = 20
            for col in ws_l.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws_l.column_dimensions[col_letter].width = max(max_len + 3, 13)
                
        excel_data_leader = io.BytesIO(); wb_leader.save(excel_data_leader); excel_data_leader.seek(0)

        # ==========================================
        # 3. 导出：完美 71 列中英文物理表头凭证接口大件
        # ==========================================
        clean_m_str = month_label.replace("月", "") 
        excel_cm_v = build_openpyxl_voucher_strict_71(df_pivot, "CM", clean_m_str)
        excel_mh_v = build_openpyxl_voucher_strict_71(df_pivot, "MH", clean_m_str)

        # ==========================================
        # 4. 前端交互呈现
        # ==========================================
        st.markdown("---")
        st.markdown("### 📥 基础业务与高管总表下载")
        c_btn1, c_btn2 = st.columns(2)
        with c_btn1:
            st.download_button(
                label=f"📊 点击下载：2026年{month_label}投放费用计提表.xlsx",
                data=excel_data_orig, file_name=f"2026年{month_label}投放费用计提表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with c_btn2:
            st.download_button(
                label=f"👑 点击下载：2026年{month_label}推广费用-各主体情况汇总.xlsx",
                data=excel_data_leader, file_name=f"2026年{month_label}推广费用-各主体情况汇总.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        st.markdown("### 💰 金蝶云星空标准财务凭证一键导出")
        c_v1, c_v2 = st.columns(2)
        with c_v1:
            st.download_button(
                label=f"🧾 点击下载：CM-投放费用金蝶上传凭证-2026{clean_m_str.zfill(2)}.xlsx",
                data=excel_cm_v, file_name=f"CM-投放费用金蝶上传凭证-2026{clean_m_str.zfill(2)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with c_v2:
            st.download_button(
                label=f"🧾 点击下载：MH-投放费用金蝶上传凭证-2026{clean_m_str.zfill(2)}.xlsx",
                data=excel_mh_v, file_name=f"MH-投放费用金蝶上传凭证-2026{clean_m_str.zfill(2)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
